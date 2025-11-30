from openpyxl import load_workbook

class SpreadsheetBuilder:
    def __init__(self, template_path: str):
        """
        Load workbook from template and prepare sheet mappings.
        """
        self.wb = load_workbook(template_path)
        self.sheet_columns = {}
        self.col_index = {}

        # Cache headers and column indices for all sheets
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            headers = [cell.value for cell in sheet[1]]  # first row
            self.sheet_columns[sheet_name] = headers
            self.col_index[sheet_name] = {h: i + 1 for i, h in enumerate(headers)}

    def add_row(self, sheet_name: str, data: dict):
        """
        Insert a new row at the bottom of the sheet.
        `data` is a dict: {column_name: value}
        """
        if sheet_name not in self.wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' does not exist in template.")

        sheet = self.wb[sheet_name]
        current_row = sheet.max_row + 1

        # Create empty row
        sheet.append([""] * len(self.sheet_columns[sheet_name]))

        # Fill values according to column names
        for col_name, value in data.items():
            if col_name not in self.col_index[sheet_name]:
                raise KeyError(f"Column '{col_name}' not found in sheet '{sheet_name}'")
            col = self.col_index[sheet_name][col_name]
            sheet.cell(row=current_row, column=col, value=value)
        

    def add_raw_data(self, raw_data: dict):
        """
        raw_data: {sheet_name: {column_name: value}}
        """
        for sheet_name, data in raw_data.items():
            self.add_row(sheet_name, data)

    def save(self, path: str):
        self.wb.save(path)
